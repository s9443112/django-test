
from msilib import type_string
from django.shortcuts import render
# from django.views.generic import ListView, DetailView
from django.http import Http404, HttpResponse
from django.core.serializers import serialize
from rest_framework import viewsets, status
from .models import Setting, Device, DeviceCount, DeviceDispatch, WorkShipDispatch, Workshop, DispatchList, User
from .models import get_Device_last_count, start_dispatch, select_dispatch_by_status, select_dispatch_by_device_limit_1, finish_dispatch, book_dispatch, search_dispatch_history, search_deviceDispatch_history, select_deviceCount_by_device, Today_dispatch
from .serializers import SettingSeializer, DeviceSeializer, DeviceCountSeializer, DeviceDispatchSeializer, WorkshopSeializer, WorkShipDispatchSeializer, DispatchListSeializer, UserSeializer
from rest_framework.decorators import action
from rest_framework.response import Response
# from rest_framework.parsers import FileUploadParser
from rest_framework.parsers import (
    FormParser,
    MultiPartParser
)
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from django.forms.models import model_to_dict
from apscheduler.schedulers.background import BackgroundScheduler
from django_apscheduler.jobstores import DjangoJobStore, register_events, register_job
from post.management.commands import export_excel
from datetime import datetime, timedelta, time
# import openpyxl
import pandas as pd
import json
import openpyxl
import os
from copy import copy


def copy_sheet(source_sheet, target_sheet):
    # copy all the cel values and styles
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min)
        # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width)  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


# class PostView(viewsets.ModelViewSet):
#     serializer_class = PostSerializer
#     queryset = Post.objects.all()


# class HistoryView(viewsets.ModelViewSet):
#     serializer_class = HistorySeializer
#     # lookup_url_kwarg = "date_time"
#     queryset = History.objects.all()

#     # Get History last limit
#     @swagger_auto_schema(
#         operation_summary='查詢歷史最後幾筆資料',
#         manual_parameters=[
#             openapi.Parameter(
#                 name='limit',
#                 in_=openapi.IN_QUERY,
#                 description='限制',
#                 type=openapi.TYPE_INTEGER
#             ),
#         ]
#     )
#     @action(detail=False, methods=['get'], url_name="Get History last limit")
#     def get_last_limit(self, request):
#         limit = int(request.query_params.get('limit', None))

#         historys = get_history_last_limit(limit=limit)
#         serializer = HistorySeializer(historys, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)

#     # Get History between DateTime
#     @swagger_auto_schema(
#         operation_summary='取得歷史藉由時間區間',

#         request_body=openapi.Schema(
#             type=openapi.TYPE_OBJECT,
#             properties={
#                 'start_time': openapi.Schema(
#                     type=openapi.TYPE_STRING,
#                     description='開始時間'
#                 ),
#                 'end_time': openapi.Schema(
#                     type=openapi.TYPE_STRING,
#                     description='結束時間'
#                 )
#             }
#         )
#     )
#     @action(detail=False, methods=['post'], url_name="Get History between DateTime")
#     def get_date_time(self, request):
#         # start_time = request.query_params.get('start_time', None)
#         # end_time = request.query_params.get('end_time', None)
#         start_time = request.data.get('start_time', None)
#         end_time = request.data.get('end_time', None)
#         historys = get_history_between_date(
#             start_time=start_time, end_time=end_time)
#         serializer = HistorySeializer(historys, many=True)
#         return Response(serializer.data, status=status.HTTP_200_OK)


class SettingView(viewsets.ModelViewSet):
    serializer_class = SettingSeializer
    queryset = Setting.objects.all()

class UserView(viewsets.ModelViewSet):
    serializer_class = UserSeializer
    queryset = User.objects.all()


class DeviceDispatchView(viewsets.ModelViewSet):
    serializer_class = DeviceDispatchSeializer
    queryset = DeviceDispatch.objects.all()

    @swagger_auto_schema(
        operation_summary='查詢所有工單by device',
        manual_parameters=[
            openapi.Parameter(
                name='user_id',
                in_=openapi.IN_QUERY,
                description='設備ID',
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                name='today_end',
                in_=openapi.IN_QUERY,
                description='結束時間',
                type=openapi.TYPE_STRING
            ),
            openapi.Parameter(
                name='today_start',
                in_=openapi.IN_QUERY,
                description='開始時間',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="DeviceDispatch by device")
    def search_DeviceDispatch_by_device(self, request):
        user_id = int(request.query_params.get('user_id', None))
        today_start = request.query_params.get('today_start', None)
        today_end = request.query_params.get('today_end', None)


        user = User.objects.filter(id=user_id)
        # device = Device.objects.filter(name=device_id)

        if user == None:
            return Response({'status': 'failed', 'msg': 'cannot find device.'}, status=status.HTTP_200_OK)

        user = User.objects.get(id=user_id)
        
        deviceDispatchList = DeviceDispatch.objects.filter(
            user=user, date_time__lte=today_end, date_time__gte=today_start)
        output = []
        for each in deviceDispatchList:
            dispatch = DispatchList.objects.get(dispatchNumber=each.dispatchNumber)
            new_obj = model_to_dict(each)
            new_obj["product_name"] = dispatch.product_name
            new_obj["material_code"] = dispatch.material_code
            output.append(new_obj)

        return Response(output, status=status.HTTP_200_OK)
        # serializer = DeviceDispatchSeializer(deviceDispatchList, many=True)
        # return Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='查詢工單by dispatchNumber',
        manual_parameters=[
            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單號碼',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="DeviceDispatch by dispatch")
    def search_deviceDispatch_history(self, request):
        dispatchNumber = request.query_params.get('dispatchnumber', None)

        dispatchlist = search_deviceDispatch_history(
            dispatchNumber=dispatchNumber)

        data = []
        print(dispatchlist)

        for each in dispatchlist:
            print(each.device.id)
            data.append({
                "device_name": each.device.name,
                "device_id": each.device.id,
                "count": each.count,
                "date_time": each.date_time
            })
        print(data)

        # serializer = DeviceDispatchSeializer(dispatchlist, many=True)
        return Response(data, status=status.HTTP_200_OK)


class DispatchListView(viewsets.ModelViewSet):
    serializer_class = DispatchListSeializer
    queryset = DispatchList.objects.all()

    @swagger_auto_schema(
        operation_summary="工單完工",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'dispatchNumber': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='工單編號'
                ),
            }
        )
    )
    @action(detail=False, methods=['post'], url_name="finish_dispatch")
    def finish_dispatch(self, request):
        dispatchNumber = request.data.get('dispatchNumber', None)

        dispatchlist = finish_dispatch(
            dispatchNumber=dispatchNumber)
        return Response(dispatchlist, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='工單派工',

        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'dispatchNumber': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='工單編號'
                ),
                'publish_count': openapi.Schema(
                    type=openapi.TYPE_NUMBER,
                    description='推播警戒值'
                ),
                'devices_first': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='設備工單'
                ),
                'devices_second': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='設備預排工單2'
                ),
                'devices_third': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='設備預排工單3'
                )
            }
        )
    )
    @action(detail=False, methods=['post'], url_name="start_dispatch")
    def start_dispatch(self, request):
        # start_time = request.query_params.get('start_time', None)
        # end_time = request.query_params.get('end_time', None)
        dispatchNumber = request.data.get('dispatchNumber', None)
        publish_count = int(request.data.get('publish_count', 0))
        devices_first = request.data.get('devices_first', None)
        devices_second = request.data.get('devices_second', None)
        devices_third = request.data.get('devices_third', None)

        dispatchlist = start_dispatch(
            dispatchNumber=dispatchNumber, devices_first=devices_first, devices_second=devices_second, devices_third=devices_third,publish_count=publish_count)

        return Response(dispatchlist, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary="下載person dispatch",

        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'json_data': openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    description='json_data'
                )
            }
        )
    )
    @action(detail=False, methods=['post'], url_name="person dispatch")
    def download_json(self, request):
        json_data = request.data.get('json_data', None)
        # print(json_data)
        workbook = openpyxl.load_workbook('個人生產月報表_example.xlsx')
        copy_sheeter = workbook['sheet1']
        copy_sheete2 = workbook['sheet2']

        workbook2 = openpyxl.Workbook()
        workbook2.create_sheet('Sheet2')
        sheet = workbook2['Sheet']
        sheet2 = workbook2['Sheet2']

        try:
            copy_sheet(copy_sheeter, sheet)
            copy_sheet(copy_sheete2, sheet2)
        except Exception as e:
            print(e)
        start_row = 4
        for each in json_data.keys():
            # print(each)
            start_col = 2
            sheet.cell(start_row, 1).value = datetime.strptime(
                each, "%Y-%m-%d")
            for dispatch in json_data[each]:
                print(dispatch)

                sheet.cell(
                    start_row, start_col).value = "{} /\n {}".format(dispatch["dispatchNumber"],dispatch["material_code"])
                sheet.cell(start_row, start_col+1).value = dispatch["count"]
                start_col = start_col + 2
            start_row = start_row + 1

        start_row = 5
        for each in json_data.keys():
            print(User.objects.get(id=int(json_data[each][0]["user"])))
            sheet2.cell(3, 1).value = ' 部門名稱:熱處理 - 剪口組                                                 (工號) 操作者: ' + \
                User.objects.get(id=int(json_data[each][0]["user"])).name
            # print(each)
            start_col = 3
            sheet2.cell(start_row, start_col -
                        1).value = datetime.strptime(each, "%Y-%m-%d")
            buff_row = start_row
            buff_col = start_col
            check = False
            for idx, dispatch in enumerate(json_data[each]):
                # print(dispatch)

                if idx > 4 and check == False:
                    check = True
                    buff_row = start_row
                    buff_col = start_col+2

                sheet2.cell(
                    buff_row, buff_col).value = "{} /\n {}".format(dispatch["dispatchNumber"],dispatch["material_code"])
                sheet2.cell(buff_row, buff_col+1).value = dispatch["count"]
                buff_row = buff_row + 1
            start_row = start_row + 3

        workbook2.save('person_dispatch.xlsx')

        with open('person_dispatch.xlsx', 'rb') as fh:
            data = fh.read()
        response = HttpResponse(
            data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=person_dispatch.xlsx'
        return response

    @swagger_auto_schema(
        operation_summary="下載today dispatch",

        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'json_data': openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    description='json_data'
                ),
                'date': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='date'
                )
            },
        )
    )
    @action(detail=False, methods=['post'], url_name="today dispatch")
    def download_today_dispatch_json(self, request):
        json_data = request.data.get('json_data', None)
        date = request.data.get('date', None)
        print(json_data)
        workbook = openpyxl.load_workbook('主管_example.xlsx')
        copy_sheeter = workbook['sheet1']

        workbook2 = openpyxl.Workbook()

        sheet = workbook2['Sheet']

        try:
            copy_sheet(copy_sheeter, sheet)
        except Exception as e:
            print(e)
        start_row = 5
        for each in json_data.keys():
            # print(each)
            start_col = 3
            sheet.cell(3, 1).value = ' 部門名稱:熱處理 - 剪口組                           生產日報表                     日期:' + date
            sheet.cell(start_row, start_col-2).value = each
            buff_row = start_row
            buff_col = start_col
            check = False
            for idx, dispatch in enumerate(json_data[each]):
                print(dispatch)
                print(idx)

                if idx > 4 and check == False:
                    check = True
                    buff_row = start_row
                    buff_col = start_col+4
                sheet.cell(
                    buff_row, buff_col).value = dispatch["material_code"]
                sheet.cell(
                    buff_row, buff_col+1).value = dispatch["dispatchNumber"]
                sheet.cell(buff_row, buff_col+2).value = dispatch["count"]
                buff_row = buff_row + 1

            start_row = start_row + 5

        workbook2.save('today_dispatch.xlsx')

        with open('today_dispatch.xlsx', 'rb') as fh:
            data = fh.read()
        response = HttpResponse(
            data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=today_dispatch.xlsx'
        return response

    @swagger_auto_schema(
        operation_summary='查詢今日執行工單',
        manual_parameters=[
            openapi.Parameter(
                name='date',
                in_=openapi.IN_QUERY,
                description='時間',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get DispatchDetail")
    def select_dispatchdetail_by_today(self, request):

        date = request.query_params.get('date', None)
        # print(date)

        dispatchlist = Today_dispatch(date=date)
        return Response(dispatchlist, status=status.HTTP_200_OK)

    # Get DispatchList by status

    @swagger_auto_schema(
        operation_summary='查詢工單by status',
        manual_parameters=[
            openapi.Parameter(
                name='status',
                in_=openapi.IN_QUERY,
                description='狀態',
                type=openapi.TYPE_INTEGER
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get Dispatch by status")
    def select_dispatch_by_status(self, request):
        _status = int(request.query_params.get('status', None))
        print(_status)

        dispatchlist = select_dispatch_by_status(status=_status)
        serializer = DispatchListSeializer(dispatchlist, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='查詢工單by dispatchNumber',
        manual_parameters=[
            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單號碼',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get Dispatch by dispatch")
    def select_dispatch_by_dispatchNumber(self, request):
        dispatchNumber = request.query_params.get('dispatchnumber', None)

        dispatchlist = DispatchList.objects.filter(
            dispatchNumber=dispatchNumber)
        print(dispatchlist)

        serializer = DispatchListSeializer(dispatchlist[0], many=False)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='查詢工單by dispatchNumber',
        manual_parameters=[
            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單號碼',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get DispatchDetail")
    def select_dispatchdetail_by_dispatchNumber(self, request):
        dispatchNumber = request.query_params.get('dispatchnumber', None)
        dispatchlist = search_dispatch_history(dispatchNumber=dispatchNumber)
        return Response(dispatchlist, status=status.HTTP_200_OK)


class DeviceView(viewsets.ModelViewSet):
    serializer_class = DeviceSeializer
    queryset = Device.objects.all()

    @swagger_auto_schema(
        operation_summary='查詢所有設備最後計數',
        manual_parameters=[
            openapi.Parameter(
                name='id',
                in_=openapi.IN_QUERY,
                description='設備ID',
                type=openapi.TYPE_INTEGER
            ),
        ]

    )
    @action(detail=False, methods=['get'], url_name="Get Device last count")
    def get_last_device_count(self, request):

        id = int(request.query_params.get('id', None))
        devicecount = get_Device_last_count(id=id)
        if devicecount == None:
            msg = {'msg': 'not exist'}

            return Response(msg, status=status.HTTP_200_OK)

        serializer = DeviceCountSeializer(devicecount, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='查詢所有設備by user',
        manual_parameters=[
            openapi.Parameter(
                name='user',
                in_=openapi.IN_QUERY,
                description='User ID',
                type=openapi.TYPE_INTEGER
            ),
        ]

    )
    @action(detail=False, methods=['get'], url_name="Get Device by user")
    def get_device_by_user(self, request):

        id = int(request.query_params.get('user', None))
        user = User.objects.get(id=id)
        

        device = Device.objects.get(user=user)

        serializer = DeviceSeializer(device, many=False)
        return Response(serializer.data, status=status.HTTP_200_OK)

    

    @swagger_auto_schema(
        operation_summary='工單預排',

        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'dispatchNumber': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='工單編號'
                ),
                'devices_second': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='設備預排工單2'
                ),
                'devices_third': openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description='設備預排工單3'
                )
            }
        )
    )
    @action(detail=False, methods=['post'], url_name="book_dispatch")
    def book_dispatch(self, request):
        dispatchNumber = request.data.get('dispatchNumber', None)

        devices_second = request.data.get('devices_second', None)
        devices_third = request.data.get('devices_third', None)

        dispatchlist = book_dispatch(
            dispatchNumber=dispatchNumber, devices_second=devices_second, devices_third=devices_third)

        return Response(dispatchlist, status=status.HTTP_200_OK)


class DeviceCountView(viewsets.ModelViewSet):
    serializer_class = DeviceCountSeializer
    queryset = DeviceCount.objects.all()

    @swagger_auto_schema(
        operation_summary='查詢deviceCount today by device',
        manual_parameters=[

            openapi.Parameter(
                name='user_id',
                in_=openapi.IN_QUERY,
                description='工單',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get deviceCount today by device")
    def select_deviceCount_by_device(self, request):
        user_id = request.query_params.get('user_id', None)

        result = select_deviceCount_by_device(user_id=user_id)

        return Response(result, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='查詢device by 工單',
        manual_parameters=[

            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="Get Dispatch by status")
    def select_device_by_dispatch(self, request):

        dispatchNumber = request.query_params.get('dispatchnumber', None)
        # print(dispatchNumber)

        result = select_dispatch_by_device_limit_1(
            dispatchNumber=dispatchNumber)
        if result == None:
            msg = {'msg': 'not exist'}

            return Response(msg, status=status.HTTP_200_OK)
        print(result)
        # serializer = DeviceCountSeializer(result, many=True)
        # qs_json = serialize('json', result)

        return Response(result, status=status.HTTP_200_OK)

        # return HttpResponse(result, content_type='application/json')


class WorkshopView(viewsets.ModelViewSet):
    serializer_class = WorkshopSeializer
    queryset = Workshop.objects.all()


class WorkShipDispatchView(viewsets.ModelViewSet):
    serializer_class = WorkShipDispatchSeializer
    queryset = WorkShipDispatch.objects.all()


class UploadDispatchFileView(viewsets.ModelViewSet):

    parser_classes = (FormParser, MultiPartParser)

    @swagger_auto_schema(
        operation_summary='下載標準表格',
        manual_parameters=[
            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單號碼',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="download by example table")
    def dowload_example(self, request):
        pass

    @swagger_auto_schema(
        operation_summary='下載已完成的工單 dispatchNumber',
        manual_parameters=[
            openapi.Parameter(
                name='dispatchnumber',
                in_=openapi.IN_QUERY,
                description='工單號碼',
                type=openapi.TYPE_STRING
            ),
        ]
    )
    @action(detail=False, methods=['get'], url_name="download by dispatch")
    def dowload_filename(self, request):
        dispatchNumber = request.query_params.get('dispatchnumber', None)

        dispatchlist = search_deviceDispatch_history(
            dispatchNumber=dispatchNumber)

        data = []

        for each in dispatchlist:
            print(each.device.id)
            data.append({
                # "生產日期": each.date_time.strftime('%Y-%m-%d'),
                # "設備名稱": each.device.name,
                # "生產數量": each.count,
                "DateTime": each.date_time.strftime('%Y-%m-%d'),
                "DeviceName": each.device.name,
                "DeviceCount": each.count,
            })
        print(data)

        df = pd.DataFrame(data)
        # writer = pd.ExcelWriter('{}工單報表.xlsx'.format(dispatchNumber), engine='xlsxwriter')
        # df.to_excel(writer, sheet_name='Sheet1', startrow=1, header=False, index=False)
        print(df)
        # results = pd.Dataframe()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = "attachment;filename*=UTF-8''" + \
            '{}.csv'.format(dispatchNumber)
        # df.to_excel(writer, sheet_name='Sheet1', startrow=1, header=False, index=Falsem)
        df.to_csv(path_or_buf=response, sep=',', float_format='%.2f',
                  index=False, decimal=",", encoding="big5")
        return response

        # return Response(data, status=status.HTTP_200_OK)

    @swagger_auto_schema(
        operation_summary='工單上傳',
        manual_parameters=[
            openapi.Parameter(
                name='filename',
                in_=openapi.IN_FORM,
                description='工單清單檔案',
                type=openapi.TYPE_FILE
            )
        ]

    )
    @action(detail=False, methods=['post'], parser_classes=(MultiPartParser, ), url_name="uploadfile")
    def upload(self, request):
        # print(request.FILES.get('filename'))
        # dispatchfile = request.data.get('filename', None)
        dispatchfile = request.data['filename']
        print(dispatchfile)

        df = pd.read_excel(dispatchfile, engine='openpyxl')

        # print(df["工單編號"])
        dispatchlist = []
        for i in df.index:
            if str(df["工單編號"][i]) == 'nan':
                continue
            # print("20{}".format(df["預計開工日"][i]))
            # print(pd.to_datetime("20{}".format(df["預計開工日"][i]),format='%Y-%m-%d'))
            dispatchlist.append(DispatchList(
                department=df["部門"][i],
                dispatchNumber=df["工單編號"][i],
                real_start_date=None,
                real_end_date=None,
                guess_start_date=pd.to_datetime("20{}".format(
                    df["預計開工日"][i]), format='%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'),
                guess_end_date=pd.to_datetime("20{}".format(
                    df["預計完工日"][i]), format='%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S'),
                generate_start_date=datetime.today().strftime('%Y-%m-%d %H:%M:%S'),
                material_code=df["料件編號"][i],
                product_name=df["品名"][i],
                specification=df["規格"][i],
                qt_count=df["生產數量"][i],
                real_count=0,
                status=0
            ))
        # print(model_to_dict(dispatchlist[0]))

        for x in dispatchlist:
            print(model_to_dict(x))
        DispatchList.objects.bulk_create(dispatchlist)

        return Response({'status': 'success', 'msg': 'ok'}, status=status.HTTP_200_OK)


# # 实例化调度器
# scheduler = BackgroundScheduler(timezone='Asia/Taipei')
# # 调度器使用默认的DjangoJobStore()
# scheduler.add_jobstore(DjangoJobStore(), 'default')

# # 每天8点半执行这个任务
# @register_job(scheduler, 'cron', id='test', hour=15, minute=7,args=['test'],replace_existing=True)
# def test(s):
#     # 具体要执行的代码
#     export_excel.Command().handle()
#     pass

# # 注册定时任务并开始
# register_events(scheduler)
# scheduler.start()
