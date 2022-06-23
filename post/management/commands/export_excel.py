from django.core.management.base import BaseCommand
from post import models
from datetime import datetime, timedelta, time
from django.forms.models import model_to_dict
import pandas as pd
from string import ascii_lowercase
import itertools
from post.lib.mail import SendMail
import random
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
def iter_all_strings():
    for size in itertools.count(1):
        for s in itertools.product(ascii_lowercase, repeat=size):
            # print(type(s[0]))
            # print(s[0])
            if s[0] == 'a' or s[0] == 'b':
                # print('ignore')
                pass
            else:
                yield "".join(s)



def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])
    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)
def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


class Command(BaseCommand):

    def __init__(self, result: str = None, writer: str = None, workbook: str = None,filename:str =None):
        self.result = result
        self.writer = writer
        self.workbook = workbook
        self.filename = filename



    def make_sheet(self, sheet_name: str, today_start: str, today_end: str):

        device_count_list = []

        print(" \
        SELECT A.id,DATE(A.date_time),COUNT(*),B.name  \
        FROM post_devicecount A \
        LEFT JOIN post_device B \
        where A.device_id = B.id and A.date_time between '{}' and '{}'  \
        GROUP BY DATE(A.date_time),A.device_id  \
        order by DATE(A.date_time) asc \
        ".format(today_start, today_end))

        for device_count in models.DeviceCount.objects.raw(" \
        SELECT A.id,DATE(A.date_time),COUNT(*),B.name  \
        FROM post_devicecount A \
        LEFT JOIN post_device B \
        where A.device_id = B.id and A.date_time between '{}' and '{}'  \
        GROUP BY DATE(A.date_time),A.device_id  \
        order by DATE(A.date_time) desc \
        ".format(today_start, today_end)):
            print(model_to_dict(device_count))
            result = next((
                x for x in device_count_list if x["date_time"] == device_count.date_time.strftime('%Y-%m-%d')), None)
            if result == None:
                device_count_list.append({
                    "date_time": device_count.date_time.strftime('%Y-%m-%d'),
                    str(device_count.device): device_count.count
                })
            else:

                result[str(device_count.device)] = device_count.count

        df = pd.DataFrame(device_count_list)

        df.to_excel(self.writer, sheet_name=sheet_name)
        worksheet = self.writer.sheets[sheet_name]

        chart = self.workbook.add_chart({'type': 'column'})
        print(len(df.keys()))
        if len(df.keys()) == 0:
            print("沒有資料")
            return
        name_length = 2
        sheet = ''
        for s in itertools.islice(iter_all_strings(), len(df.keys())-1):
            # print(s)

            chart.add_series({
                'name': [sheet_name, 0, name_length],
                'values': '={}!${}$2:${}${}'.format(sheet_name, s, s, len(device_count_list)+1),
                'categories': '={}!$B$2:$B${}'.format(sheet_name, len(device_count_list)+1)
            })
            name_length = name_length + 1
            # sheet = s

        for s in itertools.islice(iter_all_strings(), name_length):
            sheet = s

        chart.set_x_axis({'name': 'date_time'})
        print(sheet)
        sheet = sheet.upper()

        worksheet.insert_chart('{}{}'.format(sheet.upper(), 2), chart)

    def bk_today_sheet(self, result: None, writer: None, workbook: None):
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        today_start = datetime.combine(today, time())
        today_end = datetime.combine(tomorrow, time())

        device_count_list = []

        for each in result:
            print(each.name)
            device_count = models.DeviceCount.objects.filter(
                device=each, date_time__lte=today_end, date_time__gte=today_start).order_by("-id")[:1]

            if len(device_count) == 0:
                pass
            else:

                device_count_list.append({
                    'device': each.name,
                    'client_id': each.client_id,
                    'count': device_count[0].count,
                })

        # print(device_count_list)

        df = pd.DataFrame(device_count_list)

        df.to_excel(writer, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']

        chart = workbook.add_chart({'type': 'column'})
        chart.add_series({
            'values': '=Sheet1!$D$2:$D${}'.format(len(device_count_list)+1),
            'categories': '=Sheet1!$B$2:$B${}'.format(len(device_count_list)+1)
        })
        chart.set_x_axis({'name': 'device'})

        worksheet.insert_chart('F2', chart)
        # writer.save()

    def today_sheet(self):
        today = datetime.now().date()
        tomorrow = today + timedelta(1)
        today_start = datetime.combine(today, time())
        today_end = datetime.combine(tomorrow, time())

        self.make_sheet(today_start=today_start,
                        today_end=today_end, sheet_name='本日')

    def month_sheet(self):
        today = datetime.now().date() + timedelta(1)
        tomorrow = today - timedelta(30)
        today_start = datetime.combine(tomorrow, time())
        today_end = datetime.combine(today, time())

        self.make_sheet(today_start=today_start,
                        today_end=today_end, sheet_name='本月')

    def workshop_global_sheet_detail(self, result: object):
        
        rm_result = []
        for each in result:
            index = next((
                x for x, item in enumerate(rm_result) if item["workshop"] == model_to_dict(each)["workshop"]), None)
            if index == None:
                rm_result.append(model_to_dict(each))
            else:
                rm_result[index]["DeviceDispatch"].extend(model_to_dict(each)["DeviceDispatch"])

        
        # print(rm_result)
         
        for each in rm_result:
            # print(model_to_dict(each))
            workshop = models.Workshop.objects.get(
                id=each["workshop"])
            device_dispatch_list = []
            sheet_name = "{}統計日報表".format(workshop.name)
            # self.workshop_next_next(result=model_to_dict(each)["DeviceDispatch"],sheet_name="{}日報表".format(workshop.name))
            for eeach in each["DeviceDispatch"]:
                print(model_to_dict(eeach))
                result_check = next((
                    x for x in device_dispatch_list if x["date_time"] == model_to_dict(eeach)["date_time"].strftime('%Y-%m-%d')), None)
                device = models.Device.objects.get(
                    id=str(model_to_dict(eeach)["device"]))

                if result_check == None:
                    # print(model_to_dict(each.workshop))
                    device_dispatch_list.append({
                        "date_time": eeach.date_time.strftime('%Y-%m-%d'),
                        device.name: eeach.count
                    })
                else:
                    result_check[device.name] = eeach.count
            print("fuck you ass hole")
            print(device_dispatch_list)
            # print(model_to_dict(eeach))

            df = pd.DataFrame(device_dispatch_list)
            df.to_excel(self.writer, sheet_name=sheet_name)
            worksheet = self.writer.sheets[sheet_name]
            chart = self.workbook.add_chart({'type': 'column'})

            if len(df.keys()) == 0:
                print("沒有資料")
                return
            name_length = 2
            sheet = ''
            for s in itertools.islice(iter_all_strings(), len(df.keys())-1):
                # print(s)

                chart.add_series({
                    'name': [sheet_name, 0, name_length],
                    'values': '={}!${}$2:${}${}'.format(sheet_name, s, s, len(device_dispatch_list)+1),
                    'categories': '={}!$B$2:$B${}'.format(sheet_name, len(device_dispatch_list)+1)
                })
                name_length = name_length + 1
                # sheet = s

            for s in itertools.islice(iter_all_strings(), name_length):
                sheet = s

            chart.set_x_axis({'name': 'date_time'})
            print(sheet)
            sheet = sheet.upper()

            worksheet.insert_chart('{}{}'.format(sheet.upper(), 2), chart)

            # print(device_dispatch_list)

            print("==============================")
        self.writer.save()
         


        rm_result = []
        for each in result:
            index = next((
                x for x, item in enumerate(rm_result) if item["workshop"] == model_to_dict(each)["workshop"]), None)
            if index == None:
                rm_result.append(model_to_dict(each))
            else:
                rm_result[index]["DeviceDispatch"].extend(model_to_dict(each)["DeviceDispatch"])
            
        print(rm_result)
        for each in rm_result:
            
            # print(model_to_dict(each))
            workshop = models.Workshop.objects.get(
                id=each["workshop"])
            device_dispatch_list = []
            sheet_name = "{}統計日報表".format(workshop.name)
            self.workshop_next_next(result=each["DeviceDispatch"],sheet_name="{}日報表".format(workshop.name))
        self.workshop_next(result=result, sheet_name='光榮各加工站生產日報')
        

    def workshop_next_next(self,result: object, sheet_name: str):
        device_count_list = []
        for each in result:
            device = models.Device.objects.get(
                    id=str(model_to_dict(each)["device"]))
            device_count_list.append({
                "date_time": each.date_time.strftime('%Y-%m-%d'),
                "device": device.name,
                "count": each.count
            })
        
        print(device_count_list)

        df = pd.DataFrame(device_count_list)
       

        workbook = openpyxl.load_workbook('workshop.xlsx')
        # workbook.create_sheet(sheet_name)
        sheet = workbook['sheet1']
        row = 8
        for idx, val in enumerate(device_count_list):
            sheet.cell(row+idx,1).value = val["device"]
            sheet.cell(row+idx,4).value = val["count"]
            sheet.cell(row+idx,7).value = val["date_time"]
      
        workbook2 = openpyxl.load_workbook(self.filename)
        workbook2.create_sheet(sheet_name)
        target_sheet = workbook2[sheet_name]
        try:
            copy_sheet(sheet, target_sheet)
        except Exception as e:
            print(e)
        workbook2.save(self.filename)
        print(sheet_name)
        
        print("/////////////////////////////////////////////////")

    def workshop_next(self, result: object, sheet_name: str):
        device_count_list = []
        for each in result:
            device_count_list.append({
                "date_time": each.date_time.strftime('%Y-%m-%d'),
                "workshop": str(model_to_dict(each.workshop)["name"]),
                "count": each.all_count
            })
            # print(model_to_dict(each))
       
        df = pd.DataFrame(device_count_list)
       

        workbook = openpyxl.load_workbook('workshop.xlsx')
        # workbook.create_sheet(sheet_name)
        sheet = workbook['sheet1']
        row = 8
        for idx, val in enumerate(device_count_list):
            sheet.cell(row+idx,1).value = val["workshop"]
            sheet.cell(row+idx,4).value = val["count"]
            sheet.cell(row+idx,7).value = val["date_time"]

        workbook2 = openpyxl.load_workbook(self.filename)
        workbook2.create_sheet(sheet_name)
        target_sheet = workbook2[sheet_name]
        copy_sheet(sheet, target_sheet)
        workbook2.save(self.filename)


        

      

    def workshop_global_sheet_today(self, sheet_name: str):
        today = datetime.now().date() + timedelta(1)
        tomorrow = today - timedelta(30)
        today_start = datetime.combine(tomorrow, time())
        today_end = datetime.combine(today, time())
        device_count_list = []
        result = models.WorkShipDispatch.objects.filter(
            date_time__lte=today_end, date_time__gte=today_start)

        self.workshop_next(result=result, sheet_name='光榮各加工站生產日報')

        
        for each in result:
            result_check = next((
                x for x in device_count_list if x["date_time"] == each.date_time.strftime('%Y-%m-%d')), None)
            if result_check == None:
                print(model_to_dict(each.workshop))
                device_count_list.append({
                    "date_time": each.date_time.strftime('%Y-%m-%d'),
                    str(model_to_dict(each.workshop)["name"]): each.all_count
                })
            else:
                result_check[str(model_to_dict(each.workshop)
                                 ["name"])] = each.all_count

        df = pd.DataFrame(device_count_list)

        print(df)
        df.to_excel(self.writer, sheet_name=sheet_name)
        worksheet = self.writer.sheets[sheet_name]
        chart = self.workbook.add_chart({'type': 'column'})

        if len(df.keys()) == 0:
            print("沒有資料")
            return
        name_length = 2
        sheet = ''
        for s in itertools.islice(iter_all_strings(), len(df.keys())-1):
            # print(s)

            chart.add_series({
                'name': [sheet_name, 0, name_length],
                'values': '={}!${}$2:${}${}'.format(sheet_name, s, s, len(device_count_list)+1),
                'categories': '={}!$B$2:$B${}'.format(sheet_name, len(device_count_list)+1)
            })
            name_length = name_length + 1
            # sheet = s

        for s in itertools.islice(iter_all_strings(), name_length):
            sheet = s

        chart.set_x_axis({'name': 'date_time'})
        print(sheet)
        sheet = sheet.upper()

        worksheet.insert_chart('{}{}'.format(sheet.upper(), 2), chart)

        self.workshop_global_sheet_detail(result=result)

        # print(result)

    def handle(self, **options):
        
       

        today_datetime = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
       
        models.Go_dispatch(today_datetime=today_datetime)
        date = datetime.today().strftime('%Y-%m-%d')
        
        devices_history = models.Today_dispatch(date=date)
        

        json_data = {}

        for ele in devices_history:
            for elee in ele["device_dispatch_list"]:
                print(elee)
                print("-----------------------")
                if elee["user_info"]["name"] not in json_data:
                     json_data[elee["user_info"]["name"]] = []
                json_data[elee["user_info"]["name"]].append({
                    "dispatchNumber": "{} /\n {}".format(ele["dispatchNumber"],ele["material_code"]),
                    "count": elee["count"]
                })


        workbook = openpyxl.load_workbook('主管_example.xlsx')
        copy_sheeter = workbook['sheet1']

        workbook2 = openpyxl.Workbook()

        sheet = workbook2['Sheet']

        try:
            copy_sheet(copy_sheeter, sheet)
        except Exception as e:
            print(e)
        start_row = 5
        for each in json_data.keys():
            # print(each)
            start_col = 3
            sheet.cell(3, 1).value = ' 部門名稱:熱處理 - 剪口組                           生產日報表                     日期:' + date
            sheet.cell(start_row, start_col-2).value = each
            buff_row = start_row
            buff_col = start_col
            check = False
            for idx, dispatch in enumerate(json_data[each]):
                print(dispatch)
                print(idx)

                if idx > 4 and check == False:
                    check = True
                    buff_row = start_row
                    buff_col = start_col+3
                sheet.cell(
                    buff_row, buff_col).value = dispatch["dispatchNumber"]
                sheet.cell(buff_row, buff_col+1).value = dispatch["count"]
                buff_row = buff_row + 1

            start_row = start_row + 5

        workbook2.save('{}_剪口團體日報表.xlsx'.format(date))
        self.writer = pd.ExcelWriter('{}_剪口團體日報表.xlsx'.format(date), engine='xlsxwriter')

        user_2 = models.User.objects.filter(auth=2)
        user_3 = models.User.objects.filter(auth=3)
        

        email_list = []
        for each in user_3:
            email_list.append(each.email)
        for each in user_2:
            email_list.append(each.email)

        SendMail.send_file(email_list=email_list,file=self.writer,title="電子計數器統計報表",msg=None)

        # self.writer = pd.ExcelWriter('{}電子計數器報表統計.xlsx'.format(today), engine='xlsxwriter')
        # self.workbook = self.writer.book

        # self.today_sheet()
        # self.month_sheet()

        # self.writer.save()

        

        ## ================================================================##

        # self.result = models.Device.objects.all()

        # for eeach in pd.date_range(start='2021-12-28', end='2021-12-30'):

        #     for each in self.result:
        #         count = random.randint(500,1200)

        #         for time in range(0,count):
        #             print("{} count is {} {}".format(eeach,count,time))
        #             models.DeviceCount.objects.create(
        #                     device=each, count=time,date_time=eeach)

        #         print("--------------------------------------")
